"""Report generation: PDF (ReportLab) and Excel (openpyxl)"""
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from datetime import datetime


BLUE = colors.HexColor("#1e40af")
LIGHT_BLUE = colors.HexColor("#dbeafe")
DARK = colors.HexColor("#1e293b")
GRAY = colors.HexColor("#64748b")
GREEN = colors.HexColor("#166534")
LIGHT_GREEN = colors.HexColor("#dcfce7")
RED = colors.HexColor("#991b1b")
LIGHT_RED = colors.HexColor("#fee2e2")


def generate_pdf_report(project: dict, calculations: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"],
                                 fontSize=20, textColor=BLUE, spaceAfter=6)
    h1_style = ParagraphStyle("H1", parent=styles["Heading1"],
                              fontSize=14, textColor=DARK, spaceAfter=4)
    h2_style = ParagraphStyle("H2", parent=styles["Heading2"],
                              fontSize=11, textColor=BLUE, spaceAfter=3)
    body = styles["Normal"]
    small = ParagraphStyle("Small", parent=body, fontSize=8, textColor=GRAY)

    story = []
    sys_specs = project.get("system_specs", {})
    name = project.get("name", "Motor Controller Design")

    # Header
    story.append(Paragraph(f"⚡ {name}", title_style))
    story.append(Paragraph(
        f"48V PMSM Motor Controller — Hardware Design Report | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        small))
    story.append(HRFlowable(width="100%", thickness=2, color=BLUE))
    story.append(Spacer(1, 6*mm))

    # System Specs table
    story.append(Paragraph("1. System Specifications", h1_style))
    spec_data = [
        ["Parameter", "Value", "Unit"],
        ["Bus Voltage (Nominal)", str(sys_specs.get("bus_voltage", 48)), "V"],
        ["Bus Voltage (Peak)", str(sys_specs.get("peak_voltage", 60)), "V"],
        ["Power", str(sys_specs.get("power", 3000)), "W"],
        ["Max Phase Current", str(sys_specs.get("max_phase_current", 80)), "A"],
        ["PWM Frequency", str(sys_specs.get("pwm_freq_hz", 20000) // 1000), "kHz"],
        ["Ambient Temperature", str(sys_specs.get("ambient_temp_c", 30)), "°C"],
        ["Gate Drive Voltage", str(sys_specs.get("gate_drive_voltage", 12)), "V"],
    ]
    story.append(_make_table(spec_data))
    story.append(Spacer(1, 6*mm))

    # MOSFET Losses
    if "mosfet_losses" in calculations:
        ml = calculations["mosfet_losses"]
        story.append(Paragraph("2. MOSFET Loss Analysis", h1_style))
        loss_data = [
            ["Parameter", "Value", "Unit"],
            ["Conduction Loss / MOSFET", str(ml.get("conduction_loss_per_fet_w", "—")), "W"],
            ["Switching Loss / MOSFET", str(ml.get("switching_loss_per_fet_w", "—")), "W"],
            ["Total Loss / MOSFET", str(ml.get("total_loss_per_fet_w", "—")), "W"],
            [f"Total Loss ({ml.get('num_fets', 6)}× MOSFETs)", str(ml.get("total_all_fets_w", ml.get("total_all_6_fets_w", "—"))), "W"],
            ["Estimated Junction Temp", str(ml.get("junction_temp_est_c", "—")), "°C"],
            ["Switch RMS Current", str(ml.get("i_rms_switch_a", "—")), "A"],
        ]
        story.append(_make_table(loss_data))
        story.append(Spacer(1, 6*mm))

    # Gate Resistors
    if "gate_resistors" in calculations:
        gr = calculations["gate_resistors"]
        story.append(Paragraph("3. Gate Drive Design", h1_style))
        gate_data = [
            ["Component", "Value", "Note"],
            ["Rg_on (Turn-On)", f"{gr.get('rg_on_recommended_ohm', '—')} Ω", "Standard E24 value"],
            ["Rg_off (Turn-Off)", f"{gr.get('rg_off_recommended_ohm', '—')} Ω", "+ 1N4148 Schottky bypass"],
            ["Rg_bootstrap", f"{gr.get('rg_bootstrap_ohm', 10)} Ω", "Bootstrap charge limiter"],
            ["Gate Rise Time", f"{gr.get('gate_rise_time_ns', '—')} ns", "Actual achieved"],
            ["dV/dt", f"{gr.get('dv_dt_v_per_us', '—')} V/µs", "Switch node slew rate"],
        ]
        story.append(_make_table(gate_data))
        story.append(Spacer(1, 6*mm))

    # Capacitors
    if "input_capacitors" in calculations:
        ic = calculations["input_capacitors"]
        story.append(Paragraph("4. Input Capacitor Design", h1_style))
        cap_data = [
            ["Component", "Value", "Qty", "Rating"],
            ["Bulk Electrolytic", f"{ic.get('c_per_bulk_cap_uf', 100)} µF", str(ic.get("n_bulk_caps", 4)), "100V, high ripple current"],
            ["Film Capacitor", f"{ic.get('c_film_uf', 4.7)} µF", "2", "100V, within 20mm of bridge"],
            ["MLCC Ceramic", f"{ic.get('c_mlcc_nf', 100)} nF", str(ic.get("c_mlcc_qty", 6)), "100V X7R, per switch node"],
        ]
        story.append(_make_table(cap_data))
        ripple_text = f"Calculated RMS ripple current: {ic.get('i_ripple_rms_a', '—')} A | Voltage ripple: {ic.get('v_ripple_actual_v', '—')} V"
        story.append(Paragraph(ripple_text, small))
        story.append(Spacer(1, 6*mm))

    # Thermal
    if "thermal" in calculations:
        th = calculations["thermal"]
        story.append(Paragraph("5. Thermal Analysis", h1_style))
        safe = th.get("thermal_safe", True)
        color_note = GREEN if safe else RED
        bg_note = LIGHT_GREEN if safe else LIGHT_RED
        therm_data = [
            ["Parameter", "Value"],
            ["Estimated Junction Temperature", f"{th.get('t_junction_est_c', '—')} °C"],
            ["Maximum Rated Junction Temp", f"{th.get('tj_max_rated_c', '—')} °C"],
            ["Thermal Margin", f"{th.get('thermal_margin_c', '—')} °C"],
            ["Power per MOSFET", f"{th.get('p_per_fet_w', '—')} W"],
            ["Total MOSFET Dissipation", f"{th.get('p_total_6_fets_w', '—')} W"],
            ["Copper Area Required", f"{th.get('copper_area_per_fet_mm2', '—')} mm² / MOSFET"],
        ]
        story.append(_make_table(therm_data))
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(f"⚠ Thermal Status: {th.get('notes', {}).get('warning', 'OK')}", small))
        story.append(Spacer(1, 6*mm))

    # Protection
    if "protection_dividers" in calculations:
        pr = calculations["protection_dividers"]
        story.append(Paragraph("6. Protection Thresholds", h1_style))
        ocp = pr.get("ocp", {})
        ovp = pr.get("ovp", {})
        uvp = pr.get("uvp", {})
        otp = pr.get("otp", {})
        tvs = pr.get("tvs", {})
        prot_data = [
            ["Protection", "Threshold", "Response"],
            ["Over-Current (OCP)", f"{ocp.get('hw_threshold_a', '—')} A", f"< {ocp.get('hw_response_us', 1)} µs"],
            ["Over-Voltage (OVP)", f"{ovp.get('trip_voltage_v', '—')} V", "Hardware comparator"],
            ["Under-Voltage (UVP)", f"{uvp.get('trip_voltage_v', '—')} V", f"Hyst: {uvp.get('hysteresis_voltage_v', 2)} V"],
            ["Over-Temp Warning", f"{otp.get('warning_temp_c', '—')} °C", "Software flag"],
            ["Over-Temp Shutdown", f"{otp.get('shutdown_temp_c', '—')} °C", "Hardware disable"],
            ["TVS Clamp", f"{tvs.get('clamping_v', '—')} V", tvs.get("part", "—")],
        ]
        story.append(_make_table(prot_data))

    story.append(Spacer(1, 10*mm))
    story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
    story.append(Paragraph("Generated by MC Hardware Designer | Anthropic Claude-powered extraction", small))

    doc.build(story)
    return buf.getvalue()


def generate_excel_report(project: dict, calculations: dict) -> bytes:
    wb = openpyxl.Workbook()

    # ── BOM Sheet ──
    ws_bom = wb.active
    ws_bom.title = "BOM"
    _excel_bom_sheet(ws_bom, calculations, project)

    # ── Calculations Sheet ──
    ws_calc = wb.create_sheet("Calculations")
    _excel_calc_sheet(ws_calc, calculations)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_bom_sheet(ws, calculations, project=None):
    headers = ["#", "Component", "Description", "Value/Part No.", "Qty", "Rating", "Est. Cost ($)"]
    _style_header_row(ws, 1, headers, "1e40af")

    bom = _build_bom(calculations, project)
    for i, row in enumerate(bom, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(wrap_text=True)
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F1F5F9")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 12


def _excel_calc_sheet(ws, calculations):
    _style_header_row(ws, 1, ["Section", "Parameter", "Value", "Unit", "Notes"], "7c3aed")
    row = 2
    for section, data in calculations.items():
        if not isinstance(data, dict):
            continue
        section_title = section.replace("_", " ").title()
        for key, val in data.items():
            if key == "notes" or isinstance(val, (dict, list)):
                continue
            ws.cell(row=row, column=1, value=section_title)
            ws.cell(row=row, column=2, value=key.replace("_", " ").title())
            ws.cell(row=row, column=3, value=val)
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value=data.get("notes", {}).get(key, "") if isinstance(data.get("notes"), dict) else "")
            row += 1

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 28


def _build_bom(calculations, project=None) -> list:
    rows = []
    n = 1
    project = project or {}

    def _get_name(block_key, fallback_param_key, fallback_name):
        try:
            raw_data = project.get("blocks", {}).get(block_key, {}).get("raw_data", {}) or {}
            part = raw_data.get("device_info", {}).get("part_number")
            comp = raw_data.get("component_name")
            return part or comp or project.get(fallback_param_key, {}).get("component_name", fallback_name)
        except AttributeError:
            return project.get(fallback_param_key, {}).get("component_name", fallback_name)

    mosfet_name = _get_name("mosfet", "mosfet_params", "N-ch Power MOSFET")
    driver_name = _get_name("driver", "driver_params", "3-phase gate driver")

    def add(component, desc, value, qty, rating, cost):
        nonlocal n
        rows.append([n, component, desc, value, qty, rating, cost])
        n += 1

    add("MOSFET", "Power switch, N-ch", f"{mosfet_name} or equiv.", 6, "Per datasheet ratings", "—")
    add("Gate Driver", "3-phase non-isolated bootstrap", f"{driver_name} or equiv.", 1, "Per datasheet ratings", "—")

    if "gate_resistors" in calculations:
        gr = calculations["gate_resistors"]
        add("Gate Resistor (ON)", "Turn-on gate resistor", f"{gr.get('rg_on_recommended_ohm', '4.7')} Ω", 6, "0402, 0.1W", "~$0.02")
        add("Gate Resistor (OFF)", "Turn-off gate resistor + bypass diode", f"{gr.get('rg_off_recommended_ohm', '2.2')} Ω + 1N4148", 6, "0402, 0.1W", "~$0.05")

    if "input_capacitors" in calculations:
        ic = calculations["input_capacitors"]
        add("Bulk Capacitor", "Input bus bulk electrolytic", f"{ic.get('c_per_bulk_cap_uf', 100)} µF / 100V", ic.get("n_bulk_caps", 4), "Panasonic EEU-FC2A101, high ripple", "~$0.60")
        add("Film Capacitor", "Mid-freq decoupling", f"{ic.get('c_film_uf', 4.7)} µF / 100V", 2, "WIMA MKP or equiv.", "~$0.80")
        add("MLCC Decoupling", "HF switch node decoupling", f"{ic.get('c_mlcc_nf', 100)} nF / 100V X7R", 6, "0603, X7R", "~$0.05")

    if "bootstrap_cap" in calculations:
        bc = calculations["bootstrap_cap"]
        add("Bootstrap Cap", "High-side gate drive bootstrap", f"{bc.get('c_boot_recommended_nf', 220)} nF / 25V", 3, "X7R MLCC, 0603", "~$0.05")
        add("Bootstrap Diode", "Bootstrap charge Schottky", bc.get("bootstrap_diode", "B0540W"), 3, "40V, 500mA, fast", "~$0.20")

    if "shunt_resistors" in calculations:
        sr = calculations["shunt_resistors"]
        ss = sr.get("single_shunt", {})
        ts = sr.get("three_shunt", {})
        add("Shunt (Single)", "Low-side current sense (single shunt mode)", f"{ss.get('value_mohm', 1)} mΩ", 1, "4-terminal Kelvin, 1W", "~$1.20")
        add("Shunt (3-phase)", "Phase current sense (3-shunt mode)", f"{ts.get('value_mohm', 0.5)} mΩ", 3, "4-terminal Kelvin, 1W", "~$1.20")

    add("TVS Diode", "Bus overvoltage transient clamp", "SMBJ58A or P6KE62A", 2, "600W, 58V clamp", "~$0.30")
    add("NTC Thermistor", "MOSFET temperature sensing", "10 kΩ @ 25°C, B=3950", 2, "0402 NTC", "~$0.15")
    add("Reverse Polarity FET", "Input reverse polarity protection", "Si7617DN or equiv.", 1, "P-ch, 60V, 50A", "~$0.90")
    add("Common Mode Choke", "EMI filter on power input", "Wurth 744235 or equiv.", 1, "5A, >100µH CM", "~$1.50")

    return rows


def _make_table(data):
    t = Table(data, hAlign="LEFT")
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_BLUE]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    t.setStyle(style)
    return t


def _style_header_row(ws, row_num, headers, hex_color):
    fill = PatternFill("solid", fgColor=hex_color)
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
